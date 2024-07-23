# python code
import tkinter as tk
from openpyxl import *
from tkinter import messagebox
from datetime import datetime, timedelta
from requests_ntlm import HttpNtlmAuth
import requests
import time
from hashlib import sha256
import pygetwindow as gw
import threading
import gspread
from oauth2client.service_account import ServiceAccountCredentials

class LoginApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Effintrack Login")
        self.root.configure(bg='white')
        
        # Labels
        self.label_username = tk.Label(self.root, text="Username:",bg='white',fg='black')
        self.label_username.pack(pady=10)
        self.entry_username = tk.Entry(self.root,highlightthickness=1, highlightbackground="black")
        self.entry_username.pack()
        
        self.label_password = tk.Label(self.root, text="Password:",bg='white',fg='black')
        self.label_password.pack(pady=10)
        self.entry_password = tk.Entry(self.root,highlightthickness=1, highlightbackground="black", show="*")
        self.entry_password.pack(pady=5)
       
        self.var_freq = 5
        
        # Login Button
        self.login_button = tk.Button(self.root, text="Login", command=self.login,width=17,height=1,bg='orange',fg='white',activebackground='navy blue', activeforeground='white')
        self.login_button.pack(pady=10)
        
    def login(self):
        username = self.entry_username.get()
        password = self.entry_password.get()
        
        # Validate credentials
        if self.validate_credentials(username, password):
            self.root.destroy()  # Close login window
            self.open_main_app(username,self.var_freq)
        else:
            messagebox.showerror("Login Failed", "Invalid username or password")
            
            

    def fetch_password(self,key):
   

    # Path to the credentials JSON file downloaded from the Google API Console
        scope = ['https://spreadsheets.google.com/feeds',
         'https://www.googleapis.com/auth/drive']

        # Path to the credentials JSON file downloaded from the Google API Console
        credentials = {'your json creds'}



        # Create credentials object
        creds = ServiceAccountCredentials.from_json_keyfile_dict(credentials, scope)

        # Authorize the client
        client = gspread.authorize(creds)

    # Open the Google Sheets spreadsheet by its title
        spreadsheet = client.open('your file name')

    # Select a specific worksheet (if not the first sheet)
        worksheet = spreadsheet.sheet1  # This selects the first sheet, replace with your sheet name if different

    # Find all values in 'key' column
        keys_column = worksheet.col_values(1)  # Assuming 'key' is in column A (1-indexed)

    # Remove the header row if it exists
        if keys_column[0] == 'key':
            keys = keys_column[1:]
        else:
            keys = keys_column

    # Find the index of the given key in the 'key' column
        try:
            key_index = keys.index(key) + 1  # +1 because gspread is 1-indexed
        except ValueError:
            print(f"Key '{key}' not found in the Google Sheet")
            return None

    # Fetch the corresponding password ('pass') from the 'pass' column
        password = worksheet.cell(key_index, 2).value  # Assuming 'pass' is in column B (1-indexed)
        if(int(worksheet.cell(key_index, 3).value)>0): # dynamic frequency of monitoring for each employee
            self.var_freq= worksheet.cell(key_index, 3).value
        print("freq captured"+str(self.var_freq))

        return password
    
    def validate_credentials(self, username, password):
        
        
        password2 = self.fetch_password(username)

        
        if (password==(password2)):
            return 1
        else:
            return 0
        
    
    def open_main_app(self, username,var_freq):
        root = tk.Tk()
        app = EmployeeTrackerApp(root, username,var_freq)
        root.mainloop()

class EmployeeTrackerApp:
    def __init__(self, root, username,var_freq):
        self.root = root
        self.root.title("Effintrack")
        self.root.configure(bg='white')
        root.geometry("240x320")
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        
        
        ###############################################

        ###############################################
        
        self.username = username
        
        self.start_time = 0.00
        self.end_time = 0.00
        self.short_break_start = 0.00
        self.short_break_end = 0.00
        self.lunch_break_start = 0.00
        self.lunch_break_end = 0.00
        self.team_meeting_start=0.00
        self.team_meeting_end = 0.00
        self.downtime_start=0.00
        self.downtime_end=0.00
        
        self.total_work_hours = 0.00       
        self.total_short_break_hours = 0.00      
        self.total_lunch_break_hours = 0.00
        self.total_team_meeting_hours = 0.00
        self.total_downtime_hours = 0.00

        self.total_break_hours = 0.00
        self.net_work_hours = 0.00
        
        self.break_active = 0
        self.session_active = 0
        self.mon_flag=0
        self.monitor_freq = int(var_freq)*60
        self.stop_event = threading.Event()
        self.current_date = datetime.now().date()  # Track current date
        
        # Labels
        self.label_status = tk.Label(self.root, text="Click 'Clock In' to start tracking",bg='white',fg='navy blue',font=('Helvetica',11))
        self.label_status.pack(pady=10)
        
        self.label_sb = tk.Label(self.root, text="Short Break Utilized:",bg='white',fg='black',font=('Helvetica',10))
        self.label_sb.pack()
        
        self.label_lb = tk.Label(self.root, text="Lunch Break Utilized:",bg='white',fg='black',font=('Helvetica',10))
        self.label_lb.pack()
        
        # Buttons
        self.clock_in_button = tk.Button(self.root, text="Clock In", width=17,height=1,command=self.clock_in, bg="green", activebackground="black", fg="white")
        self.clock_in_button.pack(pady=2)
        
        self.short_break_button = tk.Button(self.root, text="Start Short Break",width=17,height=1, command=self.start_short_break, state=tk.DISABLED,bg='#F3F3F3')
        self.short_break_button.pack(pady=5)
        
        self.lunch_break_button = tk.Button(self.root, text="Start Lunch Break", width=17,height=1,command=self.start_lunch_break, state=tk.DISABLED,bg='#F3F3F3')
        self.lunch_break_button.pack(pady=5)
                
        self.team_meeting_button = tk.Button(self.root, text="Start Team Meeting", width=17,height=1,command=self.start_team_meeting, state=tk.DISABLED,bg='#F3F3F3')
        self.team_meeting_button.pack(pady=5)
        
        self.downtime_button = tk.Button(self.root, text="Start Downtime", width=17,height=1,command=self.start_downtime, state=tk.DISABLED,bg='#F3F3F3')
        self.downtime_button.pack(pady=5)

        
        self.clock_out_button = tk.Button(self.root, text="Clock Out",width=17,height=1, command=self.clock_out, state=tk.DISABLED)
        self.clock_out_button.pack(pady=5)
        
        print('flag1')
        
        self.monitor_thread = threading.Thread(target=self.monitor_idle_time, daemon=True)
        

        # Schedule reset at midnight
        self.schedule_reset()
    
    def get_active_window_name(self):
        active_window = gw.getActiveWindow()
        if active_window:
            
            return(active_window.title)
        else:
            return("None")
        
    
    def on_close(self):
        # Define the action to be performed when the window is closed
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            print(self.session_active)
            try:
                if(self.session_active==1):
                    self.clock_out()
                else:
                    self.mon_flag=0
                    self.monitor_freq=0
                    #self.monitor_thread.join()
                    self.stop_event.set()
                    #time.sleep(1)
                self.root.destroy()
                
                
            except Exception as e:
                print(e)
                self.root.destroy()
                
            
    def schedule_reset(self):
        # Calculate time until midnight
        now = datetime.now()
        midnight = now.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)
        delta = midnight - now
        seconds_until_midnight = delta.total_seconds()

        # Schedule the reset at midnight
        self.root.after(int(seconds_until_midnight * 1000), self.midnight_reset)

    def midnight_reset(self):
        # Reset variables at midnight
        self.current_date = datetime.now().date()
        self.start_time = 0
        self.end_time = 0
        self.break_start = 0
        self.break_end = 0

        # Reschedule for the next midnight
        self.schedule_reset()
        
    def monitor_idle_time(self):
        idle_time = 0
        print('inside monitor')
        while not self.stop_event.is_set():
            print('looping')
            mon_date= str(datetime.now().strftime("%d-%m-%Y"))
            mon_time= str(datetime.now().strftime('%H:%M:%S'))
            mon_title=self.get_active_window_name()
            print(mon_title)
            # Create workbook and sheet
            self.export_event(mon_date,mon_time,mon_title)
            time.sleep(self.monitor_freq)  # Check every second

    def update_idle_label(self, seconds):
        print("Idle Time: "+ str(seconds) +" seconds")

    def clock_in(self):
        self.check_date_change()
        self.start_time = datetime.now()
        self.label_status.config(text=f"Clocked in at {self.start_time.strftime('%H:%M:%S')}")
        self.clock_in_button.config(state=tk.DISABLED,bg="#ACFF33")
        self.clock_out_button.config(state=tk.NORMAL,bg='red',fg='white')
        self.short_break_button.config(state=tk.NORMAL)
        self.team_meeting_button.config(state=tk.NORMAL)
        self.lunch_break_button.config(state=tk.NORMAL)
        self.downtime_button.config(state=tk.NORMAL)
        self.session_active=1
        self.mon_flag=1
        self.monitor_thread.start()
        
        
    def clock_out(self):
        self.label_status.config(text=f"Please wait, data is being saved.")
        self.label_sb.config(text="■ ■ ■",fg='navy blue')
        self.label_lb.config(text="")
        self.mon_flag=0
        print('closing thread')
        self.monitor_freq=0
        print(datetime.now())
        #self.monitor_thread.join()
        self.stop_event.set()
        print(datetime.now())
        
        if(self.break_active==0):
            self.check_date_change()
            self.end_time = datetime.now()
            #self.label_status.config(text=f"Clocked out at {self.end_time.strftime('%H:%M:%S')}")
            self.clock_out_button.config(state=tk.DISABLED)
            self.short_break_button.config(state=tk.DISABLED)
            self.team_meeting_button.config(state=tk.DISABLED)
            self.lunch_break_button.config(state=tk.DISABLED)
            self.downtime_button.config(state=tk.DISABLED)
            self.session_active=0
            #self.monitor_thread.join()
        else:
            messagebox.showinfo("Break period active", f"Please end your break to clock out")
            return 0
       
       
        
        messagebox.showinfo("Work Duration", f"Total work duration: {self.net_work_hours}")
         # Export to encrypted Excel file
        try:
            self.export_to_excel()
        except Exception as ee:
            messagebox.showinfo("Attention", "Fatal error, contact admin")
            
        
        # Reset times
        self.start_time = 0
        self.end_time = 0
        self.short_break_start = 0
        self.short_break_end = 0
        self.lunch_break_start = 0
        self.lunch_break_end = 0
        
        self.label_status.config(text="Click 'Clock In' to start tracking")
        if(not self.current_date==datetime.now().date() ):
            print("Date changed")
            self.clock_in_button.config(state=tk.NORMAL)
        else:
            print("same date")
            self.clock_in_button.config(state=tk.DISABLED)
        
        messagebox.showinfo("Message", "You may close the app now.")
        
        
       

    def start_short_break(self):
        self.check_date_change()
        self.short_break_start = datetime.now()
        self.label_status.config(text=f"Break started at {self.short_break_start.strftime('%H:%M:%S')}")
        self.short_break_button.config(state=tk.NORMAL,text="End Short Break", command=self.end_short_break, bg='orange' )
        self.lunch_break_button.config(state=tk.DISABLED)
        self.downtime_button.config(state=tk.DISABLED)
        self.team_meeting_button.config(state=tk.DISABLED)
        self.break_active=1
        
    def end_short_break(self):
        self.label_status.config(text=f"Clocked in at {self.start_time.strftime('%H:%M:%S')}")
        self.check_date_change()
        self.short_break_end = datetime.now()
        self.calculate_hours()
        self.label_sb.config(text=f"Short Break Utilized: {round(self.total_short_break_hours*60,2)} mins")
        self.short_break_start = 0
        self.short_break_end = 0
        print(self.total_short_break_hours*60)
        if(self.total_short_break_hours*60 <30):
            self.short_break_button.config(state=tk.NORMAL, text = "Start Short Break", command=self.start_short_break,bg="#F3F3F3")
            
        else:
            self.short_break_button.config(state=tk.DISABLED,text='Max Limit Reached')
        if(self.total_lunch_break_hours*60 <30):
            self.lunch_break_button.config(state=tk.NORMAL,text = "Start Lunch Break", command=self.start_lunch_break)
        else:
            self.lunch_break_button.config(state=tk.DISABLED,text='Max Limit Reached')
            
        self.downtime_button.config(state=tk.NORMAL, text = "Start Downtime", command=self.start_downtime)
        self.team_meeting_button.config(state=tk.NORMAL, text = "Start Team Meeting", command=self.start_team_meeting)
        self.break_active=0
            
    def start_team_meeting(self):
        self.label_status.config(text=f"Team Meeting Started")
        self.check_date_change()
        self.team_meeting_start = datetime.now()
        self.short_break_button.config(state=tk.DISABLED)
        self.lunch_break_button.config(state=tk.DISABLED)
        self.downtime_button.config(state=tk.DISABLED)
        self.team_meeting_button.config(state=tk.NORMAL, text="End Team Meeting", command=self.end_team_meeting,bg='orange')
        self.break_active=1  
        
    def end_team_meeting(self):
        self.label_status.config(text=f"Clocked in at {self.start_time.strftime('%H:%M:%S')}")
        self.check_date_change()
        self.team_meeting_end = datetime.now()
        self.calculate_hours()
        self.team_meeting_start = 0
        self.team_meeting_end = 0
        self.team_meeting_button.config(state=tk.NORMAL, text="Start Team Meeting", command=self.start_team_meeting,bg='#F3F3F3')
        self.downtime_button.config(state=tk.NORMAL, text="Start Downtime", command=self.start_downtime)
        if(self.total_short_break_hours*60 <30):
            self.short_break_button.config(state=tk.NORMAL, text = "Start Short Break", command=self.start_short_break)
        else:
            self.short_break_button.config(state=tk.DISABLED,text='Max Limit Reached')
        if(self.total_lunch_break_hours*60 <30):
            self.lunch_break_button.config(state=tk.NORMAL,text = "Start Lunch Break", command=self.start_lunch_break)
        else:
            self.lunch_break_button.config(state=tk.DISABLED,text='Max Limit Reached')
        self.break_active=0
    
    def start_downtime(self):
        self.check_date_change()
        self.downtime_start = datetime.now()
        self.label_status.config(text=f"Downtime started")
        self.short_break_button.config(state=tk.DISABLED)
        self.lunch_break_button.config(state=tk.DISABLED)
        self.team_meeting_button.config(state=tk.DISABLED)
        self.downtime_button.config(state=tk.NORMAL,text="End Downtime",command=self.end_downtime,bg='orange')
        self.break_active=1  
        
    def end_downtime(self):
        self.label_status.config(text=f"Clocked in at {self.start_time.strftime('%H:%M:%S')}")
        self.check_date_change()
        self.downtime_end = datetime.now()
        self.calculate_hours()
        self.downtime_start = 0
        self.downtime_end = 0
        self.downtime_button.config(state=tk.NORMAL,text="Start Downtime", command=self.start_downtime,bg='#F3F3F3')
        self.team_meeting_button.config(state=tk.NORMAL)
        if(self.total_short_break_hours*60 <30):
            self.short_break_button.config(state=tk.NORMAL, text = "Start Short Break", command=self.start_short_break)
        else:
            self.short_break_button.config(state=tk.DISABLED,text='Max Limit Reached')
        if(self.total_lunch_break_hours*60 <30):
            self.lunch_break_button.config(state=tk.NORMAL,text = "Start Lunch Break", command=self.start_lunch_break)
        else:
            self.lunch_break_button.config(state=tk.DISABLED,text='Max Limit Reached')
        self.break_active=0
        
    def start_lunch_break(self):
        self.check_date_change()
        self.lunch_break_start = datetime.now()
        self.label_status.config(text=f"Break started at {self.lunch_break_start.strftime('%H:%M:%S')}")
        self.lunch_break_button.config(state=tk.NORMAL, text="End Lunch Break", command=self.end_lunch_break,bg='orange')
        self.short_break_button.config(state=tk.DISABLED)
        self.downtime_button.config(state=tk.DISABLED)
        self.team_meeting_button.config(state=tk.DISABLED)
        self.break_active=1
        
    def end_lunch_break(self):
        self.label_status.config(text=f"Clocked in at {self.start_time.strftime('%H:%M:%S')}")
        self.check_date_change()
        self.lunch_break_end = datetime.now()
        self.calculate_hours()
        self.label_lb.config(text=f"Lunch Break Utilized: {round(self.total_lunch_break_hours*60 ,2)} mins")
        self.lunch_break_start = 0
        self.lunch_break_end = 0
        print(self.total_lunch_break_hours*60)
        if(self.total_lunch_break_hours*60 <30):
            self.lunch_break_button.config(state=tk.NORMAL, text="Start Lunch Break", command=self.start_lunch_break,bg='#F3F3F3')
        else:
            self.lunch_break_button.config(state=tk.DISABLED, text="Max Limit Reached")
        if(self.total_short_break_hours*60 <30):
            self.short_break_button.config(state=tk.NORMAL, text="Start Short Break", command=self.start_short_break)
        else:
            self.short_break_button.config(state=tk.DISABLED, text="Max Limit Reached")
        self.downtime_button.config(state=tk.NORMAL, text = "Start Downtime", command=self.start_downtime)
        self.team_meeting_button.config(state=tk.NORMAL, text = "Start Team Meeting", command=self.start_team_meeting)
        self.break_active=0
    
    def check_date_change(self):
        # Check if date has changed since last action
        if datetime.now().date() != self.current_date:
            self.midnight_reset()
            
    def calculate_hours(self):
        
        print("short start "+str(self.short_break_start))
        print("short end "+str(self.short_break_end))
        print("lunch start "+str(self.lunch_break_start))
        print("lunch end "+str(self.lunch_break_end))
        self.total_work_hours = (self.end_time - self.start_time).total_seconds() / 3600.0 if self.end_time and self.start_time else self.total_work_hours
        self.total_work_hours = round(self.total_work_hours,2)
        
        self.total_short_break_hours = self.total_short_break_hours + ((self.short_break_end - self.short_break_start).total_seconds() / 3600.0) if self.short_break_end and self.short_break_start else self.total_short_break_hours
        self.total_short_break_hours = round( self.total_short_break_hours,2)
        
        self.total_team_meeting_hours = self.total_team_meeting_hours + ((self.team_meeting_end - self.team_meeting_start).total_seconds() / 3600.0) if self.team_meeting_end and self.team_meeting_start else self.total_team_meeting_hours
        self.total_team_meeting_hours = round( self.total_team_meeting_hours,2)
        
        self.total_downtime_hours = self.total_downtime_hours + ((self.downtime_end - self.downtime_start).total_seconds() / 3600.0) if self.downtime_end and self.downtime_start else self.total_downtime_hours
        self.total_downtime_hours = round( self.total_downtime_hours,2)
        
        self.total_lunch_break_hours =self.total_lunch_break_hours + ((self.lunch_break_end - self.lunch_break_start).total_seconds() / 3600.0) if self.lunch_break_end and self.lunch_break_start else self.total_lunch_break_hours
        self.total_lunch_break_hours = round(self.total_lunch_break_hours,2)
        
        self.total_break_hours = self.total_short_break_hours +  self.total_lunch_break_hours
        self.total_break_hours= round(self.total_break_hours,2)
        
        self.net_work_hours = self.total_work_hours - self.total_break_hours
        self.net_work_hours = round(self.net_work_hours,2)
        
        print("Total short break "+ str(self.total_short_break_hours*60))
        print("Total lunch break "+ str(self.total_lunch_break_hours*60))
        print("Total Work hours" + str(self.total_work_hours))
        print("Total Break hours" + str(self.total_break_hours))
        print("Net work hours "+ str(self.net_work_hours))
        print("Total team meeting" + str(self.total_team_meeting_hours))
        print("Total Downtime " + str(self.total_downtime_hours))

    def export_event(self,datex,timex,event):
        try:
            scope = ['https://spreadsheets.google.com/feeds',
         'https://www.googleapis.com/auth/drive']

        # Path to the credentials JSON file downloaded from the Google API Console
            credentials = {'your creds'}



        # Create credentials object
            creds = ServiceAccountCredentials.from_json_keyfile_dict(credentials, scope)

        # Authorize the client
            client = gspread.authorize(creds)

        # Open the Google Sheets spreadsheet by its title
            spreadsheet = client.open('your filename')

        # Select a specific worksheet
            worksheet = spreadsheet.sheet1  # This selects the first sheet, replace with your sheet name if different
            next_row = len(worksheet.get_all_values()) + 1
        # Example data to write to the sheet
            data = [
        
            str(self.username),
            str(datex),
            str(timex),
            str(event)
            
        ]

        # Update the worksheet with the data
            worksheet.insert_row(data, index=next_row)
            print('Event written successfully to Google Sheets!')
        except Exception as e:
            print(e)
            
        
    def export_to_excel(self):
        self.calculate_hours()
        try:
            scope = ['https://spreadsheets.google.com/feeds',
         'https://www.googleapis.com/auth/drive']

        # Path to the credentials JSON file downloaded from the Google API Console
            credentials = {'your creds'}



        # Create credentials object
            creds = ServiceAccountCredentials.from_json_keyfile_dict(credentials, scope)

        # Authorize the client
            client = gspread.authorize(creds)

        # Open the Google Sheets spreadsheet by its title
            spreadsheet = client.open('your file name')

        # Select a specific worksheet
            worksheet = spreadsheet.sheet1  # This selects the first sheet, replace with your sheet name if different
            next_row = len(worksheet.get_all_values()) + 1
        # Example data to write to the sheet
            data = [
            str(datetime.now().strftime("%d-%m-%Y")),
            str(self.username),
            str(self.start_time.strftime('%H:%M:%S') if self.start_time else ""),
            str(self.end_time.strftime('%H:%M:%S') if self.end_time else ""),
            str(round(self.total_work_hours, 2)),
            str(round(self.total_short_break_hours*60,2)),
            str(round(self.total_lunch_break_hours*60,2)),
            str(round(self.total_team_meeting_hours*60,2)),
            str(round(self.total_downtime_hours*60,2)),
            str(round(self.total_break_hours, 2)),
            str(round(self.net_work_hours, 2))
        ]

        # Update the worksheet with the data
            worksheet.insert_row(data, index=next_row)
            time.sleep(2)
            print('Data written successfully to Google Sheets!')
        except Exception as e:
            print(e)
            
       

if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("200x200")
    app = LoginApp(root)
    root.mainloop()
